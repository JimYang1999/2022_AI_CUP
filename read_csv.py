#TEAM_1970

import pandas as pd
import os
import csv
import openpyxl

def search_file(path): #return all file name
    file=[]
    for i in os.listdir(path):
        if i not in file:
            file.append(i)
    return file

def file_name(file):
    class_={}
    for i in file: #class
        temp_file = search_file(path + i + '/') # file_name
        class_[i] = temp_file
    return class_

### 找檔案名稱 ###
path = './AICUP_private_testing/' #圖片資料夾
file = search_file(path)
class_ = file_name(file) # {class:file_name}
### 找檔案名稱###


### 取出excel ###

df = pd.read_csv('./tag_loccoor_private.csv',encoding= 'unicode_escape') #主辦方給予的csv檔案
wb = openpyxl.load_workbook('./tag_loccoor_private_class.xlsx', data_only=True) #打開新建的excel
s1 = wb['test1'] #工作表名稱 

row = 1
excel_file_name = list((df.iloc[:,1]))
for i in excel_file_name:
    col = 1
    for key in class_.keys():
        for value in class_[key]:
            if value == i:
                s1.cell(row,col).value = i #寫入excel
                s1.cell(row,col+1).value = key
                print(f'{row} done')
                row+=1
wb.save('./tag_loccoor_private_class.xlsx') #儲存excel

### 取出excel ###