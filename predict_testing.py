#TEAM_1970

import torch
from torch.utils.data import DataLoader
from torchvision import datasets, transforms
from pathlib import Path
import os
from torchvision.io import read_image
import csv
import openpyxl

model_path = f'./model.pth' #model 位置
device = "cuda" if torch.cuda.is_available() else "cpu"
model = torch.load(model_path, map_location=torch.device('cuda'))
data_dir = "./ai_cup_testing/"#須預測之圖片資料夾
pred_class = []

idx_to_classes = {
        0: 'asparagus', 1: 'bambooshoots', 2: 'betel', 3: 'broccoli', 4: 'cauliflower', 5: 'chinesecabbage',
        6: 'chinesechives', 7: 'custardapple', 8: 'grape', 9: 'greenhouse', 10: 'greenonion', 11: 'kale',
        12: 'lemon', 13: 'lettuce', 14: 'litchi', 15: 'longan', 16: 'loofah', 17: 'mango', 18: 'onion', 19: 'others',
        20: 'papaya', 21: 'passionfruit', 22: 'pear', 23: 'pennisetum', 24: 'redbeans',25: 'roseapple', 26: 'sesbania',
        27: 'soybeans', 28: 'sunhemp', 29: 'sweetpotato', 30: 'taro', 31: 'tea', 32: 'waterbamboo'}

mean = [0.485,0.456,0.406]
std = [0.229,0.224,0.225]

test_transform = transforms.Compose([
    transforms.Resize(size=(640,640)),
    transforms.ToTensor(),
    transforms.Normalize(mean,std)])

class ImageFolderWithPaths(datasets.ImageFolder):
    def __getitem__(self, index):
        original_tuple = super(ImageFolderWithPaths, self).__getitem__(index)
        path = self.imgs[index][0]
        tuple_with_path = (original_tuple + (path,))
        return tuple_with_path


dataset = ImageFolderWithPaths(root=data_dir, transform=test_transform,target_transform=None)
data_set_data_dir = DataLoader(dataset=dataset,
batch_size=1,
num_workers=0,
shuffle=False)

torch.backends.cudnn.benchmark = False
for idx, data in enumerate(data_set_data_dir):
    img,labels,paths = data
    img = img.cuda()
    output = model(img)
    output = torch.nn.functional.softmax(output, dim=-1)
    output = torch.argmax(output, dim=-1)
    pred_class.append(output)

pred_class = torch.cat(pred_class).cpu().tolist()
file_name_list = list(map(lambda x: x[0].split('\\')[-1], dataset.imgs))
true_class=[]
for i in pred_class:
    true_class.append(idx_to_classes[i])
for i in range(len(true_class)):
    print(file_name_list[i],true_class[i])

wb = openpyxl.load_workbook('./submission.xlsx', data_only=True)
s1 = wb['submission']
row = 1
for i in range(len(true_class)):
    col = 1
    s1.cell(row,col).value = file_name_list[i]
    s1.cell(row,col+1).value = true_class[i]
    print(f'{row} done')
    row+=1
wb.save('./submission.xlsx') #儲存excel